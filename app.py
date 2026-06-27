import os
import html
import json
import re
from datetime import date, timedelta
from typing import List, Dict, Any
import threading

from fastapi import FastAPI, Request
from fastapi.responses import HTMLResponse
from openpyxl import load_workbook

EXCEL_FILE = os.environ.get("EXCEL_FILE", "AIRPORT_ZONE_CORD.xlsx" if os.path.exists("AIRPORT_ZONE_CORD.xlsx") else "AIRPORT_ZONE.xlsx")
SHEET_NAME = "AIRPORT_ZONE"
HUB = "DXB"
HUB_ZONE = "Z09"

app = FastAPI(title="Airport - Zone OD Calculator")

AIRPORT_ZONE: Dict[str, str] = {}
AIRPORTS: List[str] = []
STARTUP_ERROR = ""

HIT_COUNTER_FILE = os.environ.get("HIT_COUNTER_FILE", "hit_counter.txt")
HIT_LOCK = threading.Lock()


DEFAULT_AIRPORT_COORDS = {
    "DXB": [25.2532, 55.3657],
    "JFK": [40.6413, -73.7781],
    "LHR": [51.4700, -0.4543],
    "CDG": [49.0097, 2.5479],
    "SIN": [1.3644, 103.9915],
    "MEL": [-37.6690, 144.8410],
    "AKL": [-37.0082, 174.7850],
    "SYD": [-33.9399, 151.1753],
    "BAH": [26.2708, 50.6336],
    "MCT": [23.5933, 58.2844],
    "BOM": [19.0896, 72.8656],
    "DEL": [28.5562, 77.1000],
    "MAA": [12.9941, 80.1709],
}

# Airport coordinates are loaded from Excel during startup.
# DEFAULT_AIRPORT_COORDS is only a safe fallback for older Excel files.
AIRPORT_COORDS: Dict[str, List[float]] = dict(DEFAULT_AIRPORT_COORDS)


def get_hit_count() -> int:
    try:
        if not os.path.exists(HIT_COUNTER_FILE):
            return 0
        with open(HIT_COUNTER_FILE, "r", encoding="utf-8") as f:
            return int((f.read() or "0").strip())
    except Exception:
        return 0


def increment_hit_count() -> int:
    with HIT_LOCK:
        count = get_hit_count() + 1
        with open(HIT_COUNTER_FILE, "w", encoding="utf-8") as f:
            f.write(str(count))
        return count


def norm(v) -> str:
    return str(v or "").strip().upper()


def parse_airport_coords(row, headers: List[str]) -> List[float] | None:
    """Read airport coordinates from Excel.

    Supported formats:
      AIRPORT_COORDS = [25.2532, 55.3657]
      AIRPORT_COORDS = 25.2532, 55.3657
      LATITUDE/LONGITUDE or LAT/LON columns
    """
    coord_columns = ["AIRPORT_COORDS", "AIRPORT_COORD", "COORDINATES", "COORDS", "LAT_LONG", "LATLON"]

    for col_name in coord_columns:
        if col_name in headers:
            raw = row[headers.index(col_name)]
            if raw is None:
                continue
            nums = re.findall(r"-?\d+(?:\.\d+)?", str(raw))
            if len(nums) >= 2:
                lat = float(nums[0])
                lon = float(nums[1])
                if -90 <= lat <= 90 and -180 <= lon <= 180:
                    return [lat, lon]

    lat_columns = ["LATITUDE", "LAT"]
    lon_columns = ["LONGITUDE", "LON", "LONG", "LNG"]
    lat_idx = next((headers.index(c) for c in lat_columns if c in headers), None)
    lon_idx = next((headers.index(c) for c in lon_columns if c in headers), None)

    if lat_idx is not None and lon_idx is not None:
        try:
            lat = float(row[lat_idx])
            lon = float(row[lon_idx])
            if -90 <= lat <= 90 and -180 <= lon <= 180:
                return [lat, lon]
        except Exception:
            return None

    return None


def zone_num(z: str) -> int:
    try:
        return int(str(z).upper().replace("Z", ""))
    except Exception:
        return 0


def direction(z1: str, z2: str) -> int:
    a = zone_num(z1)
    b = zone_num(z2)
    if b > a:
        return 1
    if b < a:
        return -1
    return 0


def od_pair(o: str, d: str) -> str:
    return f"{o}{d}"


def sector_pair(s: Dict[str, Any]) -> str:
    return f"{s['origin']}{s['destination']}"


def load_airport_zone():
    global AIRPORT_ZONE, AIRPORTS, AIRPORT_COORDS, STARTUP_ERROR

    AIRPORT_ZONE = {}
    AIRPORTS = []
    AIRPORT_COORDS = dict(DEFAULT_AIRPORT_COORDS)
    STARTUP_ERROR = ""

    if not os.path.exists(EXCEL_FILE):
        STARTUP_ERROR = f"Airport zone mapping not found. Please place {EXCEL_FILE} in same folder as app.py"
        return

    wb = load_workbook(EXCEL_FILE, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        STARTUP_ERROR = f"Sheet {SHEET_NAME} not found in {EXCEL_FILE}"
        return

    ws = wb[SHEET_NAME]
    headers = [norm(c.value) for c in ws[1]]

    for required in ["AIRLINE", "ZONE", "AIRPORT"]:
        if required not in headers:
            STARTUP_ERROR = "Excel must have columns AIRLINE, ZONE, AIRPORT"
            return

    zone_idx = headers.index("ZONE")
    airport_idx = headers.index("AIRPORT")

    for row in ws.iter_rows(min_row=2, values_only=True):
        airport = norm(row[airport_idx])
        zone = norm(row[zone_idx])

        if airport and zone:
            AIRPORT_ZONE[airport] = zone

            coords = parse_airport_coords(row, headers)
            if coords:
                AIRPORT_COORDS[airport] = coords

    if HUB not in AIRPORT_ZONE:
        AIRPORT_ZONE[HUB] = HUB_ZONE

    AIRPORTS = sorted(AIRPORT_ZONE.keys())


load_airport_zone()


def get_zone(airport: str) -> str:
    airport = norm(airport)
    if airport not in AIRPORT_ZONE:
        raise ValueError(f"Airport {airport} not found in AIRPORT_ZONE Excel mapping.")
    return AIRPORT_ZONE[airport]


def default_rows():
    return [{
        "sno": 1,
        "departure_date": date.today().isoformat(),
        "origin": "",
        "destination": "",
        "travel_class": "Y",
    }]


def rows_from_itinerary_string(text: str) -> List[Dict[str, Any]]:
    clean = norm(text)
    clean = clean.replace("➜", "")
    clean = clean.replace("->", "")
    clean = clean.replace("-", "")
    clean = clean.replace(",", " ")
    clean = clean.replace("/", " ")
    clean = clean.replace("|", " ")
    clean = clean.replace("\n", " ")
    clean = clean.replace("\t", " ")
    clean = clean.replace(" ", "")

    rows = []
    if len(clean) >= 6:
        idx = 1
        for pos in range(0, len(clean), 6):
            sector = clean[pos:pos + 6]
            if len(sector) < 6:
                continue

            rows.append({
                "sno": idx,
                "departure_date": date.today().isoformat(),
                "origin": sector[:3],
                "destination": sector[3:6],
                "travel_class": "Y",
            })
            idx += 1

    return rows or default_rows()


def build_sectors(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    sectors = []

    for r in rows:
        o = norm(r.get("origin"))
        d = norm(r.get("destination"))

        if not o or not d:
            continue

        oz = get_zone(o)
        dz = get_zone(d)

        sectors.append({
            "sno": r.get("sno"),
            "date": r.get("departure_date"),
            "origin": o,
            "destination": d,
            "origin_zone": oz,
            "destination_zone": dz,
            "class": norm(r.get("travel_class")) or "Y",
        })

    return sectors


def group_by_class(sectors: List[Dict[str, Any]]) -> List[List[Dict[str, Any]]]:
    groups = []
    current = []

    for s in sectors:
        if not current:
            current.append(s)
        elif s["class"] == current[-1]["class"]:
            current.append(s)
        else:
            groups.append(current)
            current = [s]

    if current:
        groups.append(current)

    return groups


def find_side_trip_indexes(sectors: List[Dict[str, Any]]) -> set:
    side_indexes = set()

    if len(sectors) < 5:
        return side_indexes

    trip_origin = sectors[0]["origin"]

    if trip_origin == HUB:
        return side_indexes

    hub_loops = []
    i = 0

    while i < len(sectors):
        if sectors[i]["origin"] == HUB and sectors[i]["destination"] != HUB:
            start = i
            end = None

            for j in range(i + 1, len(sectors)):
                if sectors[j]["origin"] != HUB and sectors[j]["destination"] == HUB:
                    end = j
                    break

            if end is not None:
                hub_loops.append((start, end))
                i = end + 1
            else:
                i += 1
        else:
            i += 1

    if not hub_loops:
        return side_indexes

    main_loop = max(hub_loops, key=lambda pair: pair[1] - pair[0])

    for start, end in hub_loops:
        if (start, end) == main_loop:
            continue

        for idx in range(start, end + 1):
            side_indexes.add(idx)

    return side_indexes


def build_virtual_legs(sectors: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    legs = []
    i = 0

    while i < len(sectors):
        s = sectors[i]

        if (
            i + 1 < len(sectors)
            and s["destination"] == HUB
            and sectors[i + 1]["origin"] == HUB
            and s["origin"] != sectors[i + 1]["destination"]
        ):
            n = sectors[i + 1]

            legs.append({
                "origin": s["origin"],
                "destination": n["destination"],
                "origin_zone": s["origin_zone"],
                "destination_zone": n["destination_zone"],
                "class": s["class"],
                "component_sectors": [sector_pair(s), sector_pair(n)],
            })

            i += 2
        else:
            legs.append({
                "origin": s["origin"],
                "destination": s["destination"],
                "origin_zone": s["origin_zone"],
                "destination_zone": s["destination_zone"],
                "class": s["class"],
                "component_sectors": [sector_pair(s)],
            })

            i += 1

    return legs


def classify_return(ods: List[Dict[str, Any]]) -> str:
    if len(ods) <= 1:
        return "ONE WAY"

    first = ods[0]
    last = ods[-1]

    if first["origin"] == last["destination"] and first["destination"] == last["origin"]:
        return "RETURN MIRRORED JOURNEY"

    return "RETURN NON MIRRORED JOURNEY"


def make_od_from_leg(leg: Dict[str, Any], turnaround="NO", reason="Sector retained as OD."):
    return {
        "origin": leg["origin"],
        "destination": leg["destination"],
        "origin_zone": leg["origin_zone"],
        "destination_zone": leg["destination_zone"],
        "class": leg["class"],
        "component_sectors": leg["component_sectors"],
        "turnaround": turnaround,
        "reasoning": reason,
    }


def calculate_main_ods(sectors: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not sectors:
        return []

    trip_origin = sectors[0]["origin"]
    legs = build_virtual_legs(sectors)

    if all(s["origin"] != HUB and s["destination"] != HUB for s in sectors):
        airports = [sectors[0]["origin"]] + [s["destination"] for s in sectors]
        zones = [sectors[0]["origin_zone"]] + [s["destination_zone"] for s in sectors]

        non_zero_signs = [
            direction(zones[i - 1], zones[i])
            for i in range(1, len(zones))
            if direction(zones[i - 1], zones[i]) != 0
        ]

        if not non_zero_signs:
            return [
                {
                    "origin": s["origin"],
                    "destination": s["destination"],
                    "origin_zone": s["origin_zone"],
                    "destination_zone": s["destination_zone"],
                    "class": s["class"],
                    "component_sectors": [sector_pair(s)],
                    "turnaround": "NO",
                    "reasoning": "Same-zone non-hub sectors retained sector-wise.",
                }
                for s in sectors
            ]

        if non_zero_signs[0] != non_zero_signs[-1]:
            turn_index = len(airports) // 2

            return [
                {
                    "origin": airports[0],
                    "destination": airports[turn_index],
                    "origin_zone": zones[0],
                    "destination_zone": zones[turn_index],
                    "class": sectors[0]["class"],
                    "component_sectors": [sector_pair(s) for s in sectors[:turn_index]],
                    "turnaround": "YES",
                    "reasoning": "Zone direction changed; turnaround point created OD break.",
                },
                {
                    "origin": airports[turn_index],
                    "destination": airports[-1],
                    "origin_zone": zones[turn_index],
                    "destination_zone": zones[-1],
                    "class": sectors[0]["class"],
                    "component_sectors": [sector_pair(s) for s in sectors[turn_index:]],
                    "turnaround": "YES",
                    "reasoning": "Inbound journey after turnaround point.",
                },
            ]

        return [{
            "origin": airports[0],
            "destination": airports[-1],
            "origin_zone": zones[0],
            "destination_zone": zones[-1],
            "class": sectors[0]["class"],
            "component_sectors": [sector_pair(s) for s in sectors],
            "turnaround": "NO",
            "reasoning": "Continuous non-hub direction; sectors combined into one OD.",
        }]

    if trip_origin == HUB:
        pure_hub_loops = False

        if len(sectors) % 2 == 0:
            pure_hub_loops = True

            for i in range(0, len(sectors), 2):
                if i + 1 >= len(sectors):
                    pure_hub_loops = False
                    break

                outbound = sectors[i]
                inbound = sectors[i + 1]

                if not (
                    outbound["origin"] == HUB
                    and outbound["destination"] != HUB
                    and inbound["origin"] != HUB
                    and inbound["destination"] == HUB
                ):
                    pure_hub_loops = False
                    break

        if pure_hub_loops:
            return [
                {
                    "origin": s["origin"],
                    "destination": s["destination"],
                    "origin_zone": s["origin_zone"],
                    "destination_zone": s["destination_zone"],
                    "class": s["class"],
                    "component_sectors": [sector_pair(s)],
                    "turnaround": "YES" if len(sectors) > 1 else "NO",
                    "reasoning": "Pure repeated hub loops retained sector based.",
                }
                for s in sectors
            ]

    if trip_origin != HUB and len(legs) == 2:
        return [
            make_od_from_leg(
                l,
                "YES",
                "Hub transfer eliminated; return journey split into outbound and inbound OD.",
            )
            for l in legs
        ]

    ods = []
    start = legs[0]

    current_origin = start["origin"]
    current_origin_zone = start["origin_zone"]
    current_class = start["class"]
    current_components = []

    prev_sign = 0
    last_dest = start["destination"]
    last_dest_zone = start["destination_zone"]

    for idx, leg in enumerate(legs):
        sign = direction(leg["origin_zone"], leg["destination_zone"])

        if idx == 0:
            prev_sign = sign
            current_components.extend(leg["component_sectors"])
            last_dest = leg["destination"]
            last_dest_zone = leg["destination_zone"]
            continue

        if prev_sign != 0 and sign != 0 and sign != prev_sign:
            ods.append({
                "origin": current_origin,
                "destination": last_dest,
                "origin_zone": current_origin_zone,
                "destination_zone": last_dest_zone,
                "class": current_class,
                "component_sectors": current_components,
                "turnaround": "YES",
                "reasoning": "Zone direction changed; turnaround point created OD break.",
            })

            current_origin = leg["origin"]
            current_origin_zone = leg["origin_zone"]
            current_class = leg["class"]
            current_components = list(leg["component_sectors"])
        else:
            current_components.extend(leg["component_sectors"])

        if sign != 0:
            prev_sign = sign

        last_dest = leg["destination"]
        last_dest_zone = leg["destination_zone"]

    ods.append({
        "origin": current_origin,
        "destination": last_dest,
        "origin_zone": current_origin_zone,
        "destination_zone": last_dest_zone,
        "class": current_class,
        "component_sectors": current_components,
        "turnaround": "NO" if len(ods) == 0 else "YES",
        "reasoning": "Continuous zone sequence; sectors combined into one OD.",
    })

    return ods


def calculate_group(group: List[Dict[str, Any]], subgroup: str) -> List[Dict[str, Any]]:
    side_indexes = find_side_trip_indexes(group)

    main_sectors = []
    side_sectors = []

    for idx, s in enumerate(group):
        if idx in side_indexes:
            side_sectors.append(s)
        else:
            main_sectors.append(s)

    main_ods = calculate_main_ods(main_sectors)
    main_classification = classify_return(main_ods)

    results = []

    for od in main_ods:
        results.append({
            "subgroup": subgroup,
            "class": od["class"],
            "od_pair": od_pair(od["origin"], od["destination"]),
            "origin_zone": od["origin_zone"],
            "destination_zone": od["destination_zone"],
            "turnaround": od["turnaround"],
            "classification": main_classification,
            "component_sectors": ", ".join(od["component_sectors"]),
            "reasoning": od["reasoning"],
        })

    for s in side_sectors:
        results.append({
            "subgroup": subgroup,
            "class": s["class"],
            "od_pair": od_pair(s["origin"], s["destination"]),
            "origin_zone": s["origin_zone"],
            "destination_zone": s["destination_zone"],
            "turnaround": "YES",
            "classification": "SIDE TRIP",
            "component_sectors": sector_pair(s),
            "reasoning": "Return journey exists inside another return itinerary; sector marked as side trip.",
        })

    return results


def calculate_all(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    sectors = build_sectors(rows)
    groups = group_by_class(sectors)

    final = []
    sno = 1

    for i, group in enumerate(groups, start=1):
        group_result = calculate_group(group, f"Group {i}")

        for r in group_result:
            r["sno"] = sno
            final.append(r)
            sno += 1

    return final


def parse_rows_from_form(form) -> List[Dict[str, Any]]:
    row_count = int(form.get("row_count", 1))
    rows = []

    for i in range(1, row_count + 1):
        rows.append({
            "sno": i,
            "departure_date": form.get(f"departure_date_{i}", date.today().isoformat()),
            "origin": norm(form.get(f"origin_{i}", "")),
            "destination": norm(form.get(f"destination_{i}", "")),
            "travel_class": norm(form.get(f"travel_class_{i}", "Y")),
        })

    return rows


def airport_select(name: str, selected: str) -> str:
    options = ['<option value=""></option>']

    for a in AIRPORTS:
        sel = "selected" if a == selected else ""
        options.append(f'<option value="{a}" {sel}>{a}</option>')

    return f'<select name="{name}">{"".join(options)}</select>'


def class_select(name: str, selected: str) -> str:
    options = []
    for c in ["F", "J", "W", "Y"]:
        sel = "selected" if c == selected else ""
        options.append(f'<option value="{c}" {sel}>{c}</option>')
    return f'<select name="{name}">{"".join(options)}</select>'


def _normal_od_css_color(index: int) -> str:
    od_colors = [
        "c-od1",
        "c-od2",
        "c-od3",
        "c-od4",
        "c-od5",
        "c-od6",
        "c-od7",
    ]
    return od_colors[min(index, len(od_colors) - 1)]


def _route_css_color(r: Dict[str, Any], normal_color_index_ref: List[int]) -> str:
    classification = str(r.get("classification", "") or "").upper()
    reasoning = str(r.get("reasoning", "") or "").upper()

    if "SIDE TRIP" in classification:
        return "c-side"
    if "ZIG-ZAG" in reasoning:
        return "c-zigzag"

    # RESTORED: each calculated OD gets its own sequential OND colour.
    # Do not group by subgroup, otherwise multiple ODs appear in the same colour.
    color = _normal_od_css_color(normal_color_index_ref[0])
    normal_color_index_ref[0] += 1
    return color

def result_color_map(results: List[Dict[str, Any]]) -> Dict[str, str]:
    colors = {}
    normal_color_index_ref = [0]

    for r in results or []:
        color = _route_css_color(r, normal_color_index_ref)
        component_text = str(r.get("component_sectors", "") or "")

        for sector in component_text.split(","):
            sector = sector.strip().replace(" ", "").upper()
            if sector:
                # Keep first OD colour for route node display, matching the earlier version.
                # Map itself stores colour per OD route, so duplicate inbound/outbound sectors
                # are still drawn with their respective OD colours.
                colors.setdefault(sector, color)

        od = str(r.get("od_pair", "") or "").strip().replace(" ", "").upper()
        if od:
            colors.setdefault(od, color)

    return colors

def build_map_routes(results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    color_map = {
        "c-preview": "#6c757d",
        "c-od1": "#e60039",
        "c-od2": "#ffb000",
        "c-od3": "#00a884",
        "c-od4": "#0077ff",
        "c-od5": "#00b8a9",
        "c-od6": "#7b2cbf",
        "c-od7": "#9c6644",
        "c-side": "#8e2de2",
        "c-zigzag": "#ff6a00",
    }

    # MAP IMPORTANT:
    # One calculated OD is drawn as one visual route group, but the route follows
    # the physical component sectors so it touches DXB and every transfer point.
    # Example: JFKDXB + DXBSIN + SINMEL is one OD-colour route group,
    # displayed through JFK -> DXB -> SIN -> MEL with direction arrows.
    normal_color_index_ref = [0]
    routes = []

    for r in results or []:
        od = str(r.get("od_pair", "") or "").strip().replace(" ", "").upper()

        if len(od) != 6:
            continue

        o = od[:3]
        d = od[3:6]

        if o not in AIRPORT_COORDS or d not in AIRPORT_COORDS:
            continue

        css_color = _route_css_color(r, normal_color_index_ref)
        component_text = str(r.get("component_sectors", "") or "")
        segments = []
        route_airports = []

        for sector in component_text.split(","):
            sector = sector.strip().replace(" ", "").upper()
            if len(sector) != 6:
                continue

            so = sector[:3]
            sd = sector[3:6]

            if so not in AIRPORT_COORDS or sd not in AIRPORT_COORDS:
                continue

            segments.append({
                "sector": sector,
                "origin": so,
                "destination": sd,
                "from": AIRPORT_COORDS[so],
                "to": AIRPORT_COORDS[sd],
            })

            if not route_airports:
                route_airports.append(so)
            route_airports.append(sd)

        if not segments:
            segments = [{
                "sector": od,
                "origin": o,
                "destination": d,
                "from": AIRPORT_COORDS[o],
                "to": AIRPORT_COORDS[d],
            }]
            route_airports = [o, d]

        route_no = len(routes) + 1
        routes.append({
            "sector": od,
            "origin": o,
            "destination": d,
            "from": AIRPORT_COORDS[o],
            "to": AIRPORT_COORDS[d],
            "segments": segments,
            "route_airports": route_airports,
            "color": color_map.get(css_color, "#6c757d"),
            "css_color": css_color,
            "od": od,
            "classification": r.get("classification", ""),
            "component_sectors": r.get("component_sectors", ""),
            "subgroup": str(r.get("subgroup", "") or "OD"),
            "route_no": route_no,
            "direction_label": "OUTBOUND" if route_no % 2 == 1 else "INBOUND",
        })

    return routes

def render_nodes(rows: List[Dict[str, Any]], results: List[Dict[str, Any]]) -> str:
    valid_rows = [r for r in rows if r.get("origin") and r.get("destination")]

    if not valid_rows:
        return '<div class="route-empty">Select origin and destination to view route nodes.</div>'

    colors = result_color_map(results)
    parts = []

    for r in valid_rows:
        sector = f"{r['origin']}{r['destination']}".replace(" ", "").upper()
        color = colors.get(sector, "c-preview")

        parts.append(
            f'''
            <span class="sector-node-wrap">
                <span class="airport-node {color}">{html.escape(r["origin"])}</span>
                <span class="route-arrow">➜</span>
                <span class="airport-node {color}">{html.escape(r["destination"])}</span>
            </span>
            '''
        )

    return "".join(parts)


def render_results(results: List[Dict[str, Any]], message: str = "") -> str:
    if message:
        return f'<span class="error">{html.escape(message)}</span>'

    if not results:
        return "No results yet."

    rows_html = ""

    for r in results:
        rows_html += f"""
        <tr>
            <td>{r.get("sno", "")}</td>
            <td>{html.escape(str(r.get("subgroup", "")))}</td>
            <td>{html.escape(str(r.get("class", "")))}</td>
            <td class="tag">{html.escape(str(r.get("od_pair", "")))}</td>
            <td>{html.escape(str(r.get("origin_zone", "")))}</td>
            <td>{html.escape(str(r.get("destination_zone", "")))}</td>
            <td>{html.escape(str(r.get("turnaround", "")))}</td>
            <td>{html.escape(str(r.get("classification", "")))}</td>
            <td>{html.escape(str(r.get("component_sectors", "")))}</td>
            <td>{html.escape(str(r.get("reasoning", "")))}</td>
        </tr>
        """

    return f"""
    <div class="table-wrap">
    <table class="result-table">
    <thead>
    <tr>
        <th>Sno</th>
        <th>Subgroup</th>
        <th>Class</th>
        <th>OD Pair</th>
        <th>Origin Zone</th>
        <th>Destination Zone</th>
        <th>Turnaround</th>
        <th>Classification</th>
        <th>Component Sectors</th>
        <th>Reasoning</th>
    </tr>
    </thead>
    <tbody>{rows_html}</tbody>
    </table>
    </div>
    """


def render_page(rows: List[Dict[str, Any]], results=None, error_message: str = "", hit_count: int = None) -> HTMLResponse:
    results = results or []

    if hit_count is None:
        hit_count = get_hit_count()

    map_routes_json = json.dumps(build_map_routes(results))
    airport_coords_json = json.dumps(AIRPORT_COORDS)

    startup_html = ""
    if STARTUP_ERROR:
        startup_html = f'<div class="error-box">{html.escape(STARTUP_ERROR)}</div>'

    row_html = ""

    for idx, r in enumerate(rows, start=1):
        dep_date = html.escape(str(r.get("departure_date") or date.today().isoformat()))
        origin = norm(r.get("origin"))
        destination = norm(r.get("destination"))
        travel_class = norm(r.get("travel_class")) or "Y"

        row_html += f"""
        <tr>
            <td><input name="sno_{idx}" value="{idx}" readonly></td>
            <td><input type="date" name="departure_date_{idx}" value="{dep_date}"></td>
            <td>{airport_select(f"origin_{idx}", origin)}</td>
            <td>{airport_select(f"destination_{idx}", destination)}</td>
            <td>{class_select(f"travel_class_{idx}", travel_class)}</td>
            <td>
                <button type="submit" name="action" value="delete_{idx}" class="clear-btn">
                    Delete
                </button>
            </td>
        </tr>
        """

    html_page = f"""
<!DOCTYPE html>
<html>
<head>
<title>EK Loyalty OD Calculator</title>

<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>

<style>
body {{ margin:0; font-family:Arial,sans-serif; background:#f7f5f2; }}
header {{
    background:linear-gradient(90deg,#7a0019,#b8860b);
    color:white;
    padding:24px 40px;
}}
.header-wrap {{ width:100%; }}
.header-top {{
    display:flex;
    justify-content:space-between;
    align-items:center;
    gap:20px;
}}
.header-title h2 {{
    margin:0;
    font-size:38px;
    font-weight:700;
}}
.hit-counter {{
    margin-left:auto;
    background:rgba(255,255,255,.18);
    border:1px solid rgba(255,255,255,.35);
    color:white;
    padding:10px 14px;
    border-radius:999px;
    font-weight:bold;
    font-size:13px;
    white-space:nowrap;
}}
.container {{ padding:24px; }}
.card {{ background:white; border-radius:14px; padding:20px; margin-bottom:22px; box-shadow:0 4px 16px rgba(0,0,0,.08); }}
.error-box {{ background:#ffe0e0; color:#990000; padding:14px; margin:20px; border-radius:10px; font-weight:bold; }}
.table-wrap {{ overflow-x:auto; }}
table {{ width:100%; border-collapse:collapse; font-size:14px; }}
th {{ background:#7a0019; color:white; padding:10px; white-space:nowrap; }}
td {{ padding:8px; border-bottom:1px solid #ddd; vertical-align:middle; }}
input,select {{ width:100%; padding:8px; border:1px solid #bbb; border-radius:6px; box-sizing:border-box; text-transform:uppercase; }}
button {{ background:#7a0019; color:white; border:none; padding:11px 18px; border-radius:8px; cursor:pointer; font-weight:bold; margin-top:14px; margin-right:8px; }}
.gold-btn {{ background:#b8860b; }}
.clear-btn {{ background:#c00020; }}
.result-table th {{ background:#b8860b; }}
.tag {{ font-weight:bold; color:#7a0019; white-space:nowrap; }}
.error {{ color:red; font-weight:bold; }}

#tableBody td:first-child, thead th:first-child {{
    width:55px; min-width:55px; max-width:55px; text-align:center;
}}
#tableBody td:first-child input {{ width:45px; text-align:center; padding:6px; }}
#tableBody td:last-child {{ width:90px; text-align:center; vertical-align:middle; }}
#tableBody td:last-child button {{
    background:#c00020;
    margin-top:0!important;
    margin-right:0;
    padding:8px 12px;
    display:inline-flex;
    align-items:center;
    justify-content:center;
}}

.itin-load-box {{
    display:flex;
    gap:8px;
    margin-bottom:14px;
    align-items:center;
}}
.itin-load-box input {{ flex:1; }}
.itin-load-box button {{
    background:#c00020;
    margin-top:0;
    white-space:nowrap;
}}

.route-flow {{
    display:flex;
    flex-wrap:wrap;
    align-items:center;
    gap:8px;
    padding:4px 0 8px 0;
    background:transparent;
    border:none;
}}
.sector-node-wrap {{
    display:inline-flex;
    align-items:center;
    gap:8px;
    margin-right:14px;
    margin-bottom:8px;
}}
.airport-node {{
    color:white;
    font-weight:bold;
    border-radius:50%;
    width:46px;
    height:46px;
    min-width:46px;
    display:inline-flex;
    align-items:center;
    justify-content:center;
    box-shadow:0 4px 10px rgba(0,0,0,.25);
    font-size:13px;
    border:2px solid white;
}}
.route-arrow {{
    font-size:22px;
    font-weight:bold;
    color:#7a0019;
}}
.legend {{
    margin-bottom:14px;
    display:flex;
    flex-wrap:wrap;
    gap:8px;
    font-size:12px;
}}
.legend-item {{
    display:inline-flex;
    align-items:center;
    gap:5px;
    background:#fafafa;
    border-radius:999px;
    padding:4px 9px;
    border:1px solid #ddd;
}}
.legend-dot {{
    width:12px;
    height:12px;
    border-radius:50%;
}}
.c-preview {{ background:#6c757d; }}
.c-od1 {{ background:#e60039; }}
.c-od2 {{ background:#ffb000; color:#222; }}
.c-od3 {{ background:#00a884; }}
.c-od4 {{ background:#0077ff; }}
.c-od5 {{ background:#00b8a9; }}
.c-od6 {{ background:#7b2cbf; }}
.c-od7 {{ background:#9c6644; }}
.c-side {{ background:#8e2de2; }}
.c-zigzag {{ background:#ff6a00; }}
.route-empty {{
    color:#777;
    font-size:13px;
    padding:4px 0;
}}

.map-popup {{
    display:none;
    position:fixed;
    inset:0;
    background:rgba(0,0,0,.65);
    z-index:9999;
    align-items:center;
    justify-content:center;
}}
.map-box {{
    width:92%;
    height:82%;
    background:white;
    border-radius:16px;
    overflow:hidden;
    box-shadow:0 10px 40px rgba(0,0,0,.4);
    position:relative;
}}
.map-header {{
    background:linear-gradient(90deg,#7a0019,#b8860b);
    color:white;
    padding:14px 18px;
    font-weight:bold;
    display:flex;
    justify-content:space-between;
    align-items:center;
}}
.close-map {{
    background:#c00020;
    color:white;
    border:none;
    border-radius:8px;
    padding:8px 12px;
    cursor:pointer;
    margin:0;
}}
#worldMap {{
    width:100%;
    height:calc(100% - 54px);
}}
</style>
</head>

<body>

{startup_html}

<header>
    <div class="header-wrap">
        <div class="header-top">
            <div class="header-title">
                <h2>Loyalty Origin and Destination (OnD) Engine</h2>
            </div>
            <div class="hit-counter">📊 Hits : {hit_count}</div>
        </div>
        <div><b>By: Syed Abbas Rizvi</b></div>
    </div>
</header>

<div class="container">

<div class="card">
<h2>Itinerary Input</h2>

<div class="route-flow">
{render_nodes(rows, results)}
</div>

<div class="legend">
    <span class="legend-item"><span class="legend-dot c-preview"></span>Draft</span>
    <span class="legend-item"><span class="legend-dot c-od1"></span>OND 1</span>
    <span class="legend-item"><span class="legend-dot c-od2"></span>OND 2</span>
    <span class="legend-item"><span class="legend-dot c-od3"></span>OND 3</span>
    <span class="legend-item"><span class="legend-dot c-od4"></span>OND 4</span>
    <span class="legend-item"><span class="legend-dot c-od5"></span>OND 5</span>
    <span class="legend-item"><span class="legend-dot c-side"></span>Side Trip</span>
    <span class="legend-item"><span class="legend-dot c-zigzag"></span>Zig-Zag</span>
</div>

<form method="post" action="/submit">
<input type="hidden" name="row_count" value="{len(rows)}">

<div class="itin-load-box">
    <input
        type="text"
        name="itinerary_string"
        placeholder="Paste itinerary string e.g. DXBSIN SINMEL MELDXB DXBLHR"
    >
    <button type="submit" name="action" value="load_string">
        Load Itinerary
    </button>
</div>

<div class="table-wrap">
<table>
<thead>
<tr>
<th style="width:55px">Sno</th>
<th>Departure Date</th>
<th>Origin</th>
<th>Destination</th>
<th>Class</th>
<th style="width:90px">Action</th>
</tr>
</thead>

<tbody id="tableBody">
{row_html}
</tbody>
</table>
</div>

<button class="gold-btn" type="submit" name="action" value="add">Add Row</button>
<button type="submit" name="action" value="calculate">Calculate OD</button>
<button class="gold-btn" type="button" onclick="openMapPopup()">View OD World Map</button>
<button class="clear-btn" type="submit" name="action" value="clear">Clear</button>
</form>
</div>

<div class="card">
<h2>Calculated OD Results</h2>
<div id="results">{render_results(results, error_message)}</div>
</div>

</div>

<div class="map-popup" id="mapPopup">
    <div class="map-box">
        <div class="map-header">
            <span>OD Route World Map</span>
            <button class="close-map" onclick="closeMapPopup()">Close</button>
        </div>
        <div id="worldMap"></div>
    </div>
</div>

<script>
const MAP_ROUTES = {map_routes_json};
const AIRPORT_COORDS_JS = {airport_coords_json};
let odMap = null;
let routeLegendControl = null;

function pairKey(a, b) {{
    return [a, b].sort().join("-");
}}

function routePairUseIndex(route) {{
    const key = pairKey(route.origin, route.destination);
    const uses = MAP_ROUTES
        .map((r, idx) => ({{ idx: idx, route_no: r.route_no || idx + 1, key: pairKey(r.origin, r.destination) }}))
        .filter(x => x.key === key)
        .sort((a, b) => a.route_no - b.route_no);

    const pos = uses.findIndex(x => x.route_no === (route.route_no || 0));
    return {{ index: pos < 0 ? 0 : pos, count: Math.max(uses.length, 1) }};
}}

function segmentPairUseIndex(segment, route) {{
    const key = pairKey(segment.origin, segment.destination);
    const uses = [];

    MAP_ROUTES.forEach((r) => {{
        (r.segments || []).forEach((seg, segIdx) => {{
            if (pairKey(seg.origin, seg.destination) === key) {{
                uses.push({{
                    route_no: r.route_no || 0,
                    segIdx: segIdx,
                    sector: seg.sector,
                    origin: seg.origin,
                    destination: seg.destination
                }});
            }}
        }});
    }});

    uses.sort((a, b) => {{
        if (a.route_no !== b.route_no) return a.route_no - b.route_no;
        if (a.segIdx !== b.segIdx) return a.segIdx - b.segIdx;
        return a.sector.localeCompare(b.sector);
    }});

    const pos = uses.findIndex(u =>
        u.route_no === (route.route_no || 0) &&
        u.sector === segment.sector &&
        u.origin === segment.origin &&
        u.destination === segment.destination
    );

    return {{ index: pos < 0 ? 0 : pos, count: Math.max(uses.length, 1) }};
}}

function curveSideForRoute(route) {{
    const use = routePairUseIndex(route);

    if (use.count > 1) {{
        return use.index % 2 === 0 ? -1 : 1;
    }}

    return (route.direction_label || "OUTBOUND") === "OUTBOUND" ? -1 : 1;
}}

function curveSideForSegment(segment, route) {{
    const use = segmentPairUseIndex(segment, route);

    if (use.count > 1) {{
        return use.index % 2 === 0 ? -1 : 1;
    }}

    return (route.direction_label || "OUTBOUND") === "OUTBOUND" ? -1 : 1;
}}

function curvedRoutePoints(from, to, curveSide, curveRatio) {{
    const lat1 = from[0];
    const lng1 = from[1];
    const lat2 = to[0];
    const lng2 = to[1];

    const dx = lng2 - lng1;
    const dy = lat2 - lat1;
    const distance = Math.sqrt(dx * dx + dy * dy) || 1;

    const curveHeight = Math.min(Math.max(distance * (curveRatio || 0.16), 2.5), 20) * curveSide;
    const midLat = (lat1 + lat2) / 2;
    const midLng = (lng1 + lng2) / 2;
    const controlLat = midLat + (-dx / distance) * curveHeight;
    const controlLng = midLng + (dy / distance) * curveHeight;

    const points = [];
    for (let i = 0; i <= 72; i++) {{
        const t = i / 72;
        const u = 1 - t;
        points.push([
            u * u * lat1 + 2 * u * t * controlLat + t * t * lat2,
            u * u * lng1 + 2 * u * t * controlLng + t * t * lng2
        ]);
    }}
    return points;
}}

function geographicCurvedRoutePoints(from, to, directionLabel, curveRatio) {{
    const lat1 = from[0];
    const lng1 = from[1];
    const lat2 = to[0];
    const lng2 = to[1];

    const dx = lng2 - lng1;
    const dy = lat2 - lat1;
    const distance = Math.sqrt(dx * dx + dy * dy) || 1;
    const midLat = (lat1 + lat2) / 2;
    const midLng = (lng1 + lng2) / 2;

    // New design rule:
    // OUTBOUND path arches geographically NORTH of the airport pair.
    // INBOUND path arches geographically SOUTH of the airport pair.
    // This avoids overlap for reverse sectors such as LHR-DXB and DXB-LHR.
    const isInbound = String(directionLabel || '').toUpperCase().includes('INBOUND');
    const northSouth = isInbound ? -1 : 1;
    const latLift = Math.min(Math.max(distance * (curveRatio || 0.22), 2.8), 24) * northSouth;

    // Small east/west offset keeps very vertical sectors visually separated too.
    const lngLift = Math.min(Math.max(distance * 0.025, 0.5), 4) * (isInbound ? -1 : 1);

    const controlLat = midLat + latLift;
    const controlLng = midLng + lngLift;

    const points = [];
    for (let i = 0; i <= 72; i++) {{
        const t = i / 72;
        const u = 1 - t;
        points.push([
            u * u * lat1 + 2 * u * t * controlLat + t * t * lat2,
            u * u * lng1 + 2 * u * t * controlLng + t * t * lng2
        ]);
    }}
    return points;
}}

function angleAt(points, idx) {{
    if (!points || points.length < 2 || !odMap) return 0;

    const p0 = points[Math.max(0, idx - 1)];
    const p1 = points[Math.min(points.length - 1, idx + 1)];
    const a = odMap.latLngToLayerPoint(L.latLng(p0[0], p0[1]));
    const b = odMap.latLngToLayerPoint(L.latLng(p1[0], p1[1]));

    return Math.atan2(b.y - a.y, b.x - a.x) * 180 / Math.PI;
}}

function addArrow(points, color, popupText, fraction, size) {{
    if (!points || points.length < 6) return;

    const idx = Math.max(2, Math.min(points.length - 3, Math.round((points.length - 1) * fraction)));
    const angle = angleAt(points, idx);
    const arrowSize = size || 20;

    L.marker(points[idx], {{
        interactive: true,
        icon: L.divIcon({{
            className: "route-arrow-icon",
            html: `
                <div style="
                    width:0;height:0;
                    border-top:${{Math.round(arrowSize * .42)}}px solid transparent;
                    border-bottom:${{Math.round(arrowSize * .42)}}px solid transparent;
                    border-left:${{arrowSize}}px solid ${{color}};
                    transform:rotate(${{angle}}deg);
                    transform-origin:center center;
                    filter:drop-shadow(0 1px 1px rgba(255,255,255,.95)) drop-shadow(0 1px 3px rgba(0,0,0,.45));
                "></div>`,
            iconSize: [arrowSize + 8, arrowSize + 8],
            iconAnchor: [(arrowSize + 8) / 2, (arrowSize + 8) / 2]
        }})
    }}).addTo(odMap).bindPopup(popupText);
}}

function addAirportMarker(code, coord, color) {{
    L.circleMarker(coord, {{
        radius: code === "DXB" ? 7 : 6,
        color: color,
        fillColor: color,
        fillOpacity: 1,
        weight: 3
    }}).addTo(odMap).bindPopup(`<b>${{code}}</b>`);

    L.marker(coord, {{
        icon: L.divIcon({{
            className: "airport-label",
            html: `<div style="font-size:12px;font-weight:bold;color:#111;background:rgba(255,255,255,.95);border:1px solid #bbb;border-radius:5px;padding:3px 7px;box-shadow:0 1px 4px rgba(0,0,0,.30);">${{code}}</div>`,
            iconSize: [46, 24],
            iconAnchor: [-8, -10]
        }})
    }}).addTo(odMap);
}}

function addRouteLegend() {{
    if (routeLegendControl) {{
        odMap.removeControl(routeLegendControl);
    }}

    routeLegendControl = L.control({{ position: "topright" }});
    routeLegendControl.onAdd = function () {{
        const div = L.DomUtil.create("div", "map-route-legend");
        let html = `
            <div style="background:rgba(255,255,255,.92);border:1px solid #ccc;border-radius:8px;padding:10px 12px;font-size:13px;line-height:1.65;box-shadow:0 2px 8px rgba(0,0,0,.18);">
                <div style="font-weight:bold;margin-bottom:4px;">OnD Path Routes</div>`;

        MAP_ROUTES.forEach((r, idx) => {{
            html += `<div><span style="display:inline-block;width:38px;border-top:5px solid ${{r.color}};margin-right:8px;vertical-align:middle;"></span>${{idx + 1}}. ${{r.od}} ${{r.direction_label || ""}}</div>`;
        }});

        html += `</div>`;
        div.innerHTML = html;
        return div;
    }};
    routeLegendControl.addTo(odMap);
}}

function openMapPopup() {{
    document.getElementById("mapPopup").style.display = "flex";

    setTimeout(() => {{
        if (!odMap) {{
            odMap = L.map("worldMap", {{ worldCopyJump: true }}).setView([20, 40], 2);

            // English-focused light basemap layer combination.
            L.tileLayer("https://{{s}}.basemaps.cartocdn.com/rastertiles/voyager_nolabels/{{z}}/{{x}}/{{y}}{{r}}.png", {{
                maxZoom: 8,
                attribution: "© OpenStreetMap contributors © CARTO"
            }}).addTo(odMap);

            L.tileLayer("https://{{s}}.basemaps.cartocdn.com/rastertiles/voyager_only_labels/{{z}}/{{x}}/{{y}}{{r}}.png", {{
                maxZoom: 8,
                attribution: ""
            }}).addTo(odMap);
        }}

        odMap.eachLayer(layer => {{
            if (!(layer instanceof L.TileLayer)) {{
                odMap.removeLayer(layer);
            }}
        }});

        addRouteLegend();

        let bounds = [];
        let markerColorByAirport = {{}};

        MAP_ROUTES.forEach((r) => {{
            const popupText = `
                <b>${{r.od}}</b><br>
                <b>${{r.direction_label || "ROUTE"}}</b>: ${{r.origin}} → ${{r.destination}}<br>
                <b>Path:</b> ${{(r.route_airports || []).join(" → ")}}<br>
                <b>Components:</b> ${{r.component_sectors}}<br>
                ${{r.classification || ""}}
            `;

            // IMPORTANT DESIGN CHANGE:
            // Do NOT draw extra direct OD lines (JFK→MEL / MEL→JFK).
            // Draw only the real inbound/outbound component path routes, solid,
            // in each OD colour, with opposite-side curves so same sectors do not overlap.
            (r.segments || []).forEach((seg) => {{
                const segPoints = geographicCurvedRoutePoints(seg.from, seg.to, r.direction_label, 0.22);

                L.polyline(segPoints, {{
                    color: r.color,
                    weight: 5,
                    opacity: 0.96,
                    smoothFactor: 1,
                    lineCap: "round",
                    lineJoin: "round"
                }}).addTo(odMap).bindPopup(popupText);

                addArrow(segPoints, r.color, popupText, 0.50, 23);
                if (segPoints.length > 20) {{
                    addArrow(segPoints, r.color, popupText, 0.72, 20);
                }}

                bounds.push(seg.from);
                bounds.push(seg.to);
                markerColorByAirport[seg.origin] = markerColorByAirport[seg.origin] || r.color;
                markerColorByAirport[seg.destination] = markerColorByAirport[seg.destination] || r.color;
            }});
        }});

        Object.keys(markerColorByAirport).forEach(code => {{
            if (AIRPORT_COORDS_JS[code]) {{
                addAirportMarker(code, AIRPORT_COORDS_JS[code], markerColorByAirport[code]);
            }}
        }});

        if (bounds.length > 0) {{
            odMap.fitBounds(bounds, {{ padding: [40, 40] }});
        }} else {{
            alert("Please calculate OD first. Map route will be available after OD calculation.");
        }}

        odMap.invalidateSize(true);
    }}, 250);
}}

function closeMapPopup() {{
    document.getElementById("mapPopup").style.display = "none";
}}
</script>

</body>
</html>
"""

    return HTMLResponse(html_page)


@app.get("/", response_class=HTMLResponse)
def home():
    hit_count = increment_hit_count()
    return render_page(default_rows(), hit_count=hit_count)


@app.post("/submit", response_class=HTMLResponse)
async def submit(request: Request):
    form = await request.form()
    action = form.get("action", "")

    if action == "clear":
        return render_page(default_rows())

    rows = parse_rows_from_form(form)

    if action == "load_string":
        itinerary_text = form.get("itinerary_string", "")
        rows = rows_from_itinerary_string(itinerary_text)

        for i, r in enumerate(rows, start=1):
            r["sno"] = i

        return render_page(rows)

    if str(action).startswith("delete_"):
        try:
            delete_idx = int(str(action).replace("delete_", ""))
            rows = [r for i, r in enumerate(rows, start=1) if i != delete_idx]
        except Exception:
            pass

        if not rows:
            rows = default_rows()

        for i, r in enumerate(rows, start=1):
            r["sno"] = i

        return render_page(rows)

    if action == "add":
        last_date = rows[-1].get("departure_date") or date.today().isoformat()
        try:
            next_date = (date.fromisoformat(last_date) + timedelta(days=1)).isoformat()
        except Exception:
            next_date = date.today().isoformat()

        rows.append({
            "sno": len(rows) + 1,
            "departure_date": next_date,
            "origin": "",
            "destination": "",
            "travel_class": "Y",
        })

        return render_page(rows)

    if action == "calculate":
        if STARTUP_ERROR:
            return render_page(rows, [], STARTUP_ERROR)

        valid_rows = [r for r in rows if r.get("origin") and r.get("destination")]

        if not valid_rows:
            return render_page(rows, [], "Please enter at least one valid itinerary row.")

        try:
            results = calculate_all(rows)
            return render_page(rows, results)
        except Exception as e:
            return render_page(rows, [], str(e))

    return render_page(rows)
